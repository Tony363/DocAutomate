#!/usr/bin/env python3
"""
Spreadsheet adapter helpers to normalize Excel/CSV submissions into markdown tables.
"""

from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path
from typing import List

from openpyxl import load_workbook  # type: ignore


def _rows_to_markdown(headers: List[str], rows: List[List[str]]) -> str:
    """Render tabular data as GitHub-flavoured markdown."""
    if not headers:
        # synthesize header based on max row length
        max_len = max((len(row) for row in rows), default=0)
        headers = [f"column_{idx+1}" for idx in range(max_len)]
    header_line = "| " + " | ".join(headers) + " |"
    separator = "| " + " | ".join(["---"] * len(headers)) + " |"
    body_lines = []
    for row in rows:
        padded = row + [""] * (len(headers) - len(row))
        body_lines.append("| " + " | ".join(padded) + " |")
    return "\n".join([header_line, separator, *body_lines])


def spreadsheet_to_markdown(path: Path) -> str:
    """
    Convert an Excel/CSV file into a markdown representation suitable for Claude ingestion.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8", errors="ignore") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        markdown = _rows_to_markdown(headers, [[cell.strip() for cell in row] for row in data_rows])
        return f"# CSV Table: {path.name}\n\n{markdown}"

    # Excel path (xlsx/xlsm/xls)
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sections = []
    for sheet in workbook.worksheets:
        header_row = []
        data_rows: List[List[str]] = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            str_row = ["" if cell is None else str(cell).strip() for cell in row]
            if row_idx == 0:
                header_row = str_row
            else:
                data_rows.append(str_row)
        markdown = _rows_to_markdown(header_row, data_rows)
        sections.append(f"## Sheet: {sheet.title}\n\n{markdown}")
    workbook.close()
    if not sections:
        return f"# Spreadsheet: {path.name}\n\n(No visible data rows detected)"
    return f"# Spreadsheet: {path.name}\n\n" + "\n\n".join(sections)
